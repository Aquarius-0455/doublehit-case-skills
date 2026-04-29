import json
import os
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SHEET_NAME = "测试用例清单"

# ─────────────────────────────────────────────────────────────
# Key 别名映射表：无论 AI 用英文还是中文变体，都归一化到标准中文 Key
# 新增别名：直接在对应列表里追加字符串即可
# ─────────────────────────────────────────────────────────────
KEY_ALIASES = {
    "用例编号": [
        "id", "case_id", "test_id", "case_no", "编号", "用例ID", "用例id",
        "序号", "no", "number", "case_number", "caseId", "testId", "testCaseId",
        "index", "序号", "sn",
    ],
    "所属模块": [
        "module", "模块", "模块名", "所属模块名", "module_name", "feature",
        "功能模块", "所在模块", "component", "sub_system", "path", "feature_area",
        "所属功能", "应用模块",
    ],
    "用例标题": [
        "title", "name", "标题", "用例名称", "测试标题", "用例名", "case_name",
        "case_title", "test_name", "测试用例标题", "名称", "summary", "test_title",
        "testCaseTitle", "test_case_name", "标题名称",
    ],
    "测试类型": [
        "type", "test_type", "类型", "测试分类", "case_type", "kind",
        "测试种类", "用例类型", "category", "testCategory", "classification",
        "test_classification", "用例分类",
    ],
    "前置条件": [
        "precondition", "preconditions", "前提条件", "预置条件", "pre_condition",
        "setup", "prerequisites", "prerequisite", "准备条件", "pre_requisite",
        "context", "setup_steps", "环境要求", "测试背景",
    ],
    "测试步骤": [
        "steps", "step", "操作步骤", "步骤", "test_steps", "test_step",
        "action", "actions", "操作", "执行步骤", "procedure", "test_actions",
        "instructions", "操作详情", "脚本步骤",
    ],
    "预期结果": [
        "expected", "expected_result", "期望结果", "预期", "result",
        "expect", "expected_results", "预期输出", "期望输出", "预计结果",
        "expected_output", "outcome", "expected_outcome", "checkpoint",
        "assertion", "验证点", "预期行为",
    ],
    "优先级": [
        "priority", "级别", "等级", "level", "prio", "pri",
        "重要程度", "优先等级", "severity", "risk", "prio_level",
        "重要性", "紧急程度",
    ],
    "备注": [
        "remark", "remarks", "note", "notes", "comment", "comments",
        "注释", "说明", "memo", "description", "desc", "extra",
        "info", "etc", "detailed_description", "补充说明", "其他信息",
    ],
}

# 构建反向查找表：任意别名 → 标准 Key（全部小写比对）
_ALIAS_LOOKUP: dict[str, str] = {}
for canonical, aliases in KEY_ALIASES.items():
    _ALIAS_LOOKUP[canonical.lower()] = canonical          # 标准 Key 本身也收录
    _ALIAS_LOOKUP[canonical] = canonical
    for alias in aliases:
        _ALIAS_LOOKUP[alias.lower()] = canonical
        _ALIAS_LOOKUP[alias] = canonical


def normalize_case_keys(case: dict) -> dict:
    """
    将 AI 生成的 JSON 各字段 Key 归一化到标准中文 Key。
    未能识别的 Key 保留原样（不丢弃），方便调试。
    """
    normalized = {}
    for k, v in case.items():
        canonical = _ALIAS_LOOKUP.get(k) or _ALIAS_LOOKUP.get(k.lower())
        if canonical:
            normalized[canonical] = v
        else:
            normalized[k] = v   # 无法映射的 Key 原样保留
    return normalized


class TestToolboxGenerator:
    """结合专业样式与自定义规则的测试用例生成器"""

    # 映射回用户定义的标准字段，并增加一列“系统补录”作为数据兜底
    HEADERS = ["用例编号", "所属模块", "用例标题", "测试类型", "前置条件", "测试步骤", "预期结果", "优先级", "备注", "系统补录"]

    # 列宽优化（最后一列给较大空间放置补录内容）
    COLUMN_WIDTHS = [12, 18, 25, 12, 25, 45, 40, 10, 15, 30]

    # 浅绿色主题样式 (C6E0B4)
    HEADER_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="000000", size=11)

    THIN_BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    def __init__(self):
        """初始化 Workbook 和 Worksheet 属性"""
        self.wb = None
        self.ws = None

    def setup_format(self):
        """初始化表头和列宽"""
        self.ws.row_dimensions[1].height = 30

        for col_idx, header in enumerate(self.HEADERS, 1):
            cell = self.ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.THIN_BORDER

        for col_idx, width in enumerate(self.COLUMN_WIDTHS, 1):
            self.ws.column_dimensions[get_column_letter(col_idx)].width = width

        self.ws.freeze_panes = "A2"

    def load_or_create(self, output_path):
        """加载现有 Workbook 或创建新的"""
        if os.path.exists(output_path):
            from openpyxl import load_workbook
            self.wb = load_workbook(output_path)
            # 按 Sheet 名精准定位，不依赖 .active
            if SHEET_NAME in self.wb.sheetnames:
                self.ws = self.wb[SHEET_NAME]
            else:
                self.ws = self.wb.active
                print(f"WARNING: Sheet '{SHEET_NAME}' not found, using active sheet: '{self.ws.title}'")
            print(f"INFO: Appending to existing file: {output_path}")
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = SHEET_NAME
            self.setup_format()
            print(f"INFO: Creating new file: {output_path}")

    def add_cases_from_json(self, cases_list):
        """将 JSON 列表数据填入 Excel（带补录机制，确保数据 0 丢失）"""
        for i, raw_case in enumerate(cases_list, 1):
            # 1. 归一化 Key（保留无法映射的原始 Key）
            case = normalize_case_keys(raw_case)

            # 2. 收集无法识别的字段数据，准备填入最后一列“系统补录”
            # 注意：排除我们已经映射到标准表头的字段
            standard_headers = self.HEADERS[:-1]
            unrecognized_data = []
            for k, v in case.items():
                # 如果这个 Key 不在标准字段范围内，且有实际值，则补录
                if k not in standard_headers and v:
                    unrecognized_data.append(f"[{k}]: {str(v)}")
            
            # 拼接补录字符串（换行显示）
            supplementary_info = "\n".join(unrecognized_data)

            # 3. 归一化后仍无匹配字段时给出明确警告
            if not any(h in case for h in standard_headers):
                print(
                    f"WARNING: 第 {i} 条用例归一化后仍无核心匹配字段！数据已全部汇总至 [系统补录] 列。"
                    f" 原始 Key: {list(raw_case.keys())}"
                )

            row_num = self.ws.max_row + 1

            # 4. 数据清洗：去掉"所属模块"开头的数字序号
            module_name = case.get("所属模块", "")
            module_name = re.sub(r'^[\d\.、\s]+', '', str(module_name))

            # 5. 构建行数据
            row_data = []
            for h in self.HEADERS:
                if h == "所属模块":
                    row_data.append(module_name)
                elif h == "系统补录":
                    row_data.append(supplementary_info)
                else:
                    row_data.append(case.get(h, ""))

            for col_idx, value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col_idx)
                cell.value = value
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    def save(self, output_path):
        """保存并打印成功信息"""
        self.wb.save(output_path)
        print(f"SUCCESS: {os.path.abspath(output_path)}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python case_exporter.py <input.json | -> [output.xlsx] [mode:new|append]")
        print("  -           : Read JSON data directly from standard input (stdin)")
        print("  mode=new    : Force create new file (overwrites existing)")
        print("  mode=append : Append to existing file (default)")
        print()
        print("Example: python case_exporter.py cases.json output.xlsx")
        sys.exit(1)

    input_arg = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else "TestCases.xlsx"
    mode = sys.argv[3].lower() if len(sys.argv) > 3 else "append"

    try:
        if input_arg == "-":
            # Read from STDIN
            print("INFO: Reading data from STDIN...")
            raw_data = sys.stdin.read()
            if not raw_data.strip():
                print("ERROR: Received empty data from STDIN")
                sys.exit(1)
            cases = json.loads(raw_data)
        else:
            # Read from File
            if not os.path.exists(input_arg):
                print(f"ERROR: File not found {input_arg}")
                sys.exit(1)
            with open(input_arg, 'r', encoding='utf-8') as f:
                cases = json.load(f)

        if not isinstance(cases, list):
            print("ERROR: JSON 文件必须是一个列表（list）")
            sys.exit(1)

        generator = TestToolboxGenerator()

        if mode == "new":
            generator.wb = Workbook()
            generator.ws = generator.wb.active
            generator.ws.title = SHEET_NAME
            generator.setup_format()
            print(f"INFO: Mode [NEW] - Creating fresh file: {output_file}")
        else:
            generator.load_or_create(output_file)

        generator.add_cases_from_json(cases)
        generator.save(output_file)

    except Exception as e:
        print(f"ERROR: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
